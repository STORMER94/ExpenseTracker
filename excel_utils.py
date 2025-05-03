import os
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from flask import send_file
from app import db
from models import Transaction, Category

def create_template():
    """Create an Excel template for transaction uploads"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Template"
    
    # Define header style
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Side(border_style="thin", color="000000")
    header_border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
    
    # Define headers
    headers = ['Amount', 'Date (YYYY-MM-DD)', 'Category ID', 'Description', 'Transaction Type']
    
    # Write headers and apply styles
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Add example row
    example_data = [
        100.00, 
        datetime.now().strftime('%Y-%m-%d'),
        1,
        'Example transaction description',
        'debit'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Add note row
    ws.cell(row=3, column=1, value="Note: Transaction Type must be either 'debit' or 'credit'")
    ws.merge_cells('A3:E3')
    note_cell = ws['A3']
    note_cell.font = Font(italic=True, color="FF0000")
    note_cell.alignment = Alignment(horizontal='left')
    
    # Adjust column widths
    column_widths = [15, 20, 15, 40, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def process_transaction_excel(file_data, user_id):
    """
    Process uploaded Excel file and create transactions
    Returns tuple: (success_count, error_list)
    """
    success_count = 0
    error_list = []
    
    try:
        # Read Excel file
        df = pd.read_excel(file_data)
        
        # Remove any unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Validate required columns
        required_columns = ['Amount', 'Date (YYYY-MM-DD)', 'Category ID', 'Transaction Type']
        for col in required_columns:
            if col not in df.columns:
                return 0, [f"Missing required column: {col}"]
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row['Amount']):
                    continue
                
                # Validate required fields
                if pd.isna(row['Date (YYYY-MM-DD)']):
                    error_list.append(f"Row {index+2}: Missing date")
                    continue
                    
                if pd.isna(row['Category ID']):
                    error_list.append(f"Row {index+2}: Missing category ID")
                    continue
                    
                if pd.isna(row['Transaction Type']):
                    error_list.append(f"Row {index+2}: Missing transaction type")
                    continue
                
                # Convert date if it's not already a datetime object
                if not isinstance(row['Date (YYYY-MM-DD)'], datetime):
                    try:
                        date = datetime.strptime(str(row['Date (YYYY-MM-DD)']), '%Y-%m-%d')
                    except ValueError:
                        error_list.append(f"Row {index+2}: Invalid date format. Use YYYY-MM-DD")
                        continue
                else:
                    date = row['Date (YYYY-MM-DD)']
                
                # Validate transaction type
                transaction_type = str(row['Transaction Type']).strip().lower()
                if transaction_type not in ['debit', 'credit']:
                    error_list.append(f"Row {index+2}: Transaction type must be 'debit' or 'credit'")
                    continue
                
                # Validate category exists and belongs to user
                category_id = int(row['Category ID'])
                category = Category.query.filter_by(id=category_id, user_id=user_id).first()
                if not category:
                    error_list.append(f"Row {index+2}: Category ID {category_id} not found or doesn't belong to you")
                    continue
                
                # Create transaction object
                amount = float(row['Amount'])
                description = str(row['Description']) if not pd.isna(row['Description']) else None
                
                transaction = Transaction(
                    amount=amount,
                    date=date,
                    transaction_type=transaction_type,
                    category_id=category_id,
                    description=description,
                    user_id=user_id,
                    fiscal_year=date.year,
                    month=date.month
                )
                
                db.session.add(transaction)
                success_count += 1
                
            except Exception as e:
                error_list.append(f"Row {index+2}: {str(e)}")
        
        # Commit all valid transactions
        if success_count > 0:
            db.session.commit()
        
        return success_count, error_list
        
    except Exception as e:
        return 0, [f"Error processing file: {str(e)}"]